import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Report')))


from sqlalchemy.orm import Session
from sqlalchemy import func
import models, database
from datetime import timedelta,datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Side, Border, Alignment
from io import BytesIO
from typing import Union
import schemas, check_rulebase
from collections import defaultdict


# openpyxl border styles
BORDER_NONE = None
BORDER_DASHDOT = 'dashDot'
BORDER_DASHDOTDOT = 'dashDotDot'
BORDER_DASHED = 'dashed'
BORDER_DOTTED = 'dotted'
BORDER_DOUBLE = 'double'
BORDER_HAIR = 'hair'
BORDER_MEDIUM = 'medium'
BORDER_MEDIUMDASHDOT = 'mediumDashDot'
BORDER_MEDIUMDASHDOTDOT = 'mediumDashDotDot'
BORDER_MEDIUMDASHED = 'mediumDashed'
BORDER_SLANTDASHDOT = 'slantDashDot'
BORDER_THICK = 'thick'
BORDER_THIN = 'thin'

def DrawLine(sheet,nRow,nCol):
    sheet.cell(row=nRow,column=nCol).border = Border(left=Side(border_style=BORDER_THIN,
                                                    color='000000'),
                                                    right=Side(border_style=BORDER_THIN,
                                                    color='000000'),
                                                    top=Side(border_style=BORDER_THIN,
                                                    color='000000'),
                                                    bottom=Side(border_style=BORDER_THIN,
                                                    color='000000'))

# Get firewall by ID
def get_firewall(db: Session, firewall_id: int):
    return db.query(models.Firewall).filter(models.Firewall.id == firewall_id).first()

# Get all firewalls
def get_firewalls(db: Session, fw_ids: Union[list[int], None], skip: int = 0, limit: int = 10):
    if fw_ids != None:
        return db.query(models.Firewall).filter(models.Firewall.id.in_(fw_ids)).all()

    return db.query(models.Firewall).all()
        

def get_rule(db: Session, rule_id: int):
    return db.query(models.Rule).filter(models.Rule.id == rule_id).first()

def get_rules(db: Session, fw_id: int):
    return db.query(models.Rule).filter(models.Rule.fw_id == fw_id).all()

def get_log(db: Session, log_id: int):
    return db.query(models.SysLog).filter(models.SysLog.id == log_id).first()

def get_weights(db: Session):
    return db.query(models.Weights).first()

def get_report(db: Session, report_id: int, type: str):
    report = db.query(models.Report).filter(models.Report.id == report_id).first()
    data = report.jsondata

    if not report: 
        return None

    if type == "security":
        return data
    elif type == "individual":
        rules = []
        for rule in data["rules"]:
            rule_types = []
            for type_name, rule_ids in data["types"].items():
                if rule["id"] in rule_ids:
                    rule_types.append(type_name)
            
            rules.append({ **rule, "types": rule_types })


        return {
            "fw_name": data["fw_name"],
            "rules": rules
        }





def get_report_history(db: Session, fw_id: int):
    reports = (
        db.query(models.Report.id, models.Report.update_ts)
        .filter(models.Report.fw_id == fw_id)
        .order_by(models.Report.update_ts.desc())
        .all()
    )

    return [
        {
            "reportId": report.id,
            "update_ts": report.update_ts.isoformat() if report.update_ts else None,
        }
        for report in reports
    ]

def get_most_recent_reports(db: Session, fw_ids: Union[list[int], None] = None):
    subquery_query = db.query(
        models.Report.fw_id,
        func.max(models.Report.update_ts).label("latest_ts")
    )

    if fw_ids is not None:
        subquery_query = subquery_query.filter(models.Report.fw_id.in_(fw_ids))

    subquery = subquery_query.group_by(models.Report.fw_id).subquery()

    results = (
        db.query(models.Report)
        .join(subquery, (models.Report.fw_id == subquery.c.fw_id) & (models.Report.update_ts == subquery.c.latest_ts))
        .all()
    )

    return results

def get_reports(argos_db: Session, report_db: Session, fw_ids: Union[list[int], None] = None):
    latest_reports = get_most_recent_reports(report_db, fw_ids)

    all_report_fw_ids = (
        report_db.query(models.Report.fw_id, models.Report.jsondata)
        .distinct(models.Report.fw_id)
        .all()
    )

    id_to_name = {
        fw_id: jsondata["fw_name"]
        for fw_id, jsondata in all_report_fw_ids
    }

    return {
        "idToName": id_to_name,
        "requestedFwData": {
            report.fw_id: {
                # "update_ts": report.update_ts.isoformat() if report.update_ts else None,
                "reportId": report.id,
                "fwName": report.jsondata["fw_name"],
                **{k: len(v) for k, v in report.jsondata["types"].items()},
            }
            for report in latest_reports
        }
    }
    



def get_report_data(db: Session, firewall_id: int = -1, fw_ids: Union[list[int], None] = None):
    ruletypes = ['expired','permanent','redundant','shadow','unused', 'unused_objects', 'dst_excessiveopen', 'port_excessiveopen', 'knownportopen', 'virusportopen', 'mgmtportopen', 'src_anyopen', 'dst_anyopen', 'noevidence', 'compliancecheck', 'disabled', 'invalid', 'manual']
    # ruletypeStrs = {'expired': 'Expired Rule', 'permanent': 'Permanent Rule', 'redundant': 'Redundant Rule', 'shadow': 'Shadow Rule', 'unused': 'Unused Rule', 'unused_objects': 'Unused Objects (Session-Based)', 'dst_excessiveopen': 'Dst Excessive Open', 'port_excessiveopen': 'Service Excessive Open', 'knownportopen': 'Well-Known Port Open', 'virusportopen': 'Virus Port Open', 'mgmtportopen': 'Mgmt Port Open', 'src_anyopen': 'Src ANY Open', 'dst_anyopen': 'Dst ANY Open', 'noevidence': 'NOEVIDENCE Rule', 'compliancecheck': 'Compliance', 'disabled': 'Inactive Rule', 'invalid': 'Invalid Rule', 'manual': 'Manual Rule'}

    # The report will include all firewalls
    if firewall_id == -1:
        all_firewalls = get_firewalls(db, None)
        firewalls = get_firewalls(db, fw_ids)

        firewallReport = {"idToName": {}, "requestedFwData": {}}

        if not firewalls:
            raise Exception("No firewall data found")
        
        for firewall in all_firewalls:
            if fw_ids is None or firewall.id in fw_ids:
                firewallData = analyze_rules(db, firewall.id)

                for ruletype in ruletypes:
                    setattr(firewallData, ruletype, len(getattr(firewallData, ruletype)))
                setattr(firewallData, "name", firewall.name)
                
                firewallReport["requestedFwData"][firewall.id] = firewallData

            firewallReport["idToName"][firewall.id] = firewall.name
        
        return firewallReport
    
    # we are generating report data for an individual firewall
    rules = get_rules(db, firewall_id)
    ruleAnalysis = analyze_rules(db, firewall_id)

    individualReport = { "fw_name": get_firewall(db, firewall_id).fw_name }
    analyzedRules = []

    for rule in rules:  # Start from row 2, as row 1 is for headers
        newRule = {}
        newRule["id"] = rule.id
        newRule["action"] = rule.action
        newRule["source"] = rule.source
        newRule["from_ip"] = rule.from_ip
        newRule["destination"] = rule.destination
        newRule["to_ip"] = rule.to_ip
        newRule["service"] = rule.service
        newRule["expire"] = rule.expire
        newRule["comment"] = rule.comment
        newRule["types"] = []

        for type, ruleSet in vars(ruleAnalysis).items():
            if len(ruleSet) > 0 and rule.id in ruleSet:
                newRule["types"].append(type)
        
        analyzedRules.append(newRule)
    
    individualReport["rules"] = analyzedRules

    return individualReport



def analyze_rules(db: Session, fw_id: int):
    rules = db.query(models.Rule).filter(models.Rule.fw_id == fw_id).all()
    analyses = db.query(models.Analyze).filter(models.Analyze.fw_id == fw_id).all()
    complianceObjects = db.query(models.ComplianceObject).filter(models.ComplianceObject.type == "wn" or models.ComplianceObject.type == "vi" or models.ComplianceObject.type == "mn").all()

    return check_rulebase.analyze(rules, analyses, complianceObjects, fw_id, db)


    # return(schemas.RuleAnalysis(fw_id='1', rules_count={}))    

def generate_report(argos_db: Session, report_db: Session, firewall_id: int):
    rules = get_rules(argos_db, firewall_id)
    ruleAnalysis = analyze_rules(argos_db, firewall_id)

    weights = get_weights()

    individualReport = { "fw_name": get_firewall(argos_db, firewall_id).fw_name }
    analyzedRules = []

    for rule in rules:  # Start from row 2, as row 1 is for headers
        newRule = {}
        newRule["id"] = rule.id
        newRule["action"] = rule.action
        newRule["source"] = rule.source
        newRule["from_ip"] = rule.from_ip
        newRule["destination"] = rule.destination
        newRule["to_ip"] = rule.to_ip
        newRule["service"] = rule.service
        newRule["expire"] = rule.expire
        newRule["comment"] = rule.comment
        
        analyzedRules.append(newRule)
    
    individualReport["rules"] = analyzedRules
    individualReport["types"] = ruleAnalysis.to_dict()
    individualReport["scores"] = ruleAnalysis.category_scores(weights, len(rules))

    return individualReport


def store_report(db: Session, report_data, firewall_id: int):
    report = models.Report(
        fw_id=firewall_id,
        jsondata=report_data,
    )
    db.add(report)
    db.commit()

    return report

def update_weights(db: Session, new_weights: dict):
    weights = db.query(models.Weights).first()
    if not weights:
        weights = models.Weights()
        db.add(weights)

    for key, value in new_weights.items():
        if hasattr(weights, key):
            setattr(weights, key, value)

    db.commit()
    return weights


def get_reports_info(db: Session):
    latest_reports = get_most_recent_reports(db)
    return {
        report.fw_id: {
            "update_ts": report.update_ts.isoformat() if report.update_ts else None,
            "report_id": report.id
        }
        for report in latest_reports
    }

        


    
def generate_report_file(db: Session):
    # Fetch firewall data from the database
    firewalls = get_firewalls(db, None)
    endCol = get_column_letter(3 + len(firewalls))

    if not firewalls:
        raise Exception("No firewall data found")
        # raise HTTPException(status_code=404, detail="No firewall data found")
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Summary"

    
    # Define headers for the Excel report
    individualHeaders = ["Rule ID", "Rule Type", "Source IP", "Source Zone", "Destination IP", "Destination Zone", "Service", "Expiration", "Comment", "Report Details"]
    
    ruletypeRows = {4: 'expired', 5: 'permanent', 6: 'redundant', 7: 'shadow', 9: 'unused', 10: 'unused_objects', 12: 'dst_excessiveopen', 13: 'port_excessiveopen', 14: 'knownportopen', 15: 'virusportopen', 16: 'mgmtportopen', 17: 'src_anyopen', 18: 'dst_anyopen', 21: 'noevidence', 22: 'compliancecheck', 30: 'disabled', 31: 'invalid', 32: 'manual'}
    ruletypeStrs = {'expired': 'Expired Rule', 'permanent': 'Permanent Rule', 'redundant': 'Redundant Rule', 'shadow': 'Shadow Rule', 'unused': 'Unused Rule', 'unused_objects': 'Unused Objects (Session-Based)', 'dst_excessiveopen': 'Dst Excessive Open', 'port_excessiveopen': 'Service Excessive Open', 'knownportopen': 'Well-Known Port Open', 'virusportopen': 'Virus Port Open', 'mgmtportopen': 'Mgmt Port Open', 'src_anyopen': 'Src ANY Open', 'dst_anyopen': 'Dst ANY Open', 'noevidence': 'NOEVIDENCE Rule', 'compliancecheck': 'Compliance', 'disabled': 'Inactive Rule', 'invalid': 'Invalid Rule', 'manual': 'Manual Rule'}
    
    alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"] = "Category"
    ws["A4"] = "Period Management"
    ws["A6"] = "Policy Utilization"
    ws["A11"] = "Policy Scope"
    ws["A14"] = "Service Safety"
    ws["A17"] = "Security Compliance"
    ws["A24"] = "Miscellaneous"
    ws["B1"] = "Details"
    ws["C1"] = "Number of Applicable Rules"
    ws["A1"].alignment = alignment
    ws["A4"].alignment = alignment
    ws["A6"].alignment = alignment
    ws["A11"].alignment = alignment
    ws["A14"].alignment = alignment
    ws["A17"].alignment = alignment
    ws["A24"].alignment = alignment
    ws["B1"].alignment = alignment
    ws["C1"].alignment = Alignment(vertical="center")
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)

    ws["B4"] = "Expired Rule"
    ws["B5"] = "Permanent Rule"
    ws["B6"] = "Redundant Rule"
    ws["B7"] = "Shadow Rule"
    ws["B8"] = "Partial Shadow Rule"
    ws["B9"] = "Unused Rule"
    ws["B10"] = "Unused Objects (Session-Based)"
    ws["B11"] = "Src Excessive Open"
    ws["B12"] = "Dst Excessive Open"
    ws["B13"] = "Service Excessive Open"
    ws["B14"] = "Well-Known Port Open"
    ws["B15"] = "Virus Port Open"
    ws["B16"] = "Mgmt Port Open"
    ws["B17"] = "Src ANY Open"
    ws["B18"] = "Dst ANY Open"
    ws["B19"] = "Service ANY Open"
    ws["B20"] = ">30 Days Past Expiration"
    ws["B21"] = "NOEVIDENCE Rule"
    ws["B22"] = "Compliance"
    ws["B23"] = "Deny Rule"
    ws["B24"] = "Check Longest Expiration Period"
    ws["B25"] = "Check Most Recent Expiration Period"
    ws["B26"] = "NORENEW Check Status"
    ws["B27"] = "Check Status"
    ws["B28"] = "Check Group-Specific Policy Registration"
    ws["B29"] = "Check Composition"
    ws["B30"] = "Inactive Rule"
    ws["B31"] = "Invalid Rule"
    ws["B32"] = "Manual Rule"
    ws["B33"] = "Unregistered Rule"
    ws["B34"] = "No Zone Rule"
    ws[f"{endCol}3"] = "Total"
    ws[f"{endCol}3"].font = Font(bold=True)


    ws.merge_cells("A1:A3")
    ws.merge_cells("A4:A5")
    ws.merge_cells("A6:A10")
    ws.merge_cells("A11:A13")
    ws.merge_cells("A14:A16")
    ws.merge_cells("A17:A23")
    ws.merge_cells("A24:A34")
    ws.merge_cells("B1:B3")
    ws.merge_cells(f"C1:{endCol}2")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions[endCol].width = 15
    

    for main_col_num, firewall in enumerate(firewalls, 3):
        rules = get_rules(db, firewall.id)

        ws.cell(row=3, column=main_col_num, value=firewall.name)
        ws.column_dimensions[get_column_letter(main_col_num)].width = 15

        ruleAnalysis = analyze_rules(db=db, fw_id=firewall.id)
        
        if rules:
            sheet = wb.create_sheet(title=f"{firewall.name}")

            curr_row = 2

            # Add headers to the first row
            for col_num, header in enumerate(individualHeaders, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = cell.font.copy(bold=True)  # Make the headers bold

            for rule in rules:  # Start from row 2, as row 1 is for headers
                start_row = curr_row

                sheet.cell(row=curr_row, column=1, value=rule.id)
                sheet.cell(row=curr_row, column=2, value=rule.action)
                sheet.cell(row=curr_row, column=3, value=rule.source)
                sheet.cell(row=curr_row, column=4, value=rule.from_ip)
                sheet.cell(row=curr_row, column=5, value=rule.destination)
                sheet.cell(row=curr_row, column=6, value=rule.to_ip)
                sheet.cell(row=curr_row, column=7, value=rule.service)
                sheet.cell(row=curr_row, column=8, value=rule.expire)
                sheet.cell(row=curr_row, column=9, value=rule.comment)

                for type, ruleSet in vars(ruleAnalysis).items():
                    if len(ruleSet) > 0 and rule.id in ruleSet:
                        sheet.cell(row=curr_row, column=10, value=ruletypeStrs[type])
                        curr_row += 1

                if curr_row == start_row:
                    curr_row += 1
                else:
                    for col in sheet.iter_cols(min_row=start_row, max_row=curr_row - 1, min_col=1, max_col=9):
                        sheet.merge_cells(start_column=col[0].column, end_column=col[0].column, start_row=start_row, end_row=curr_row-1)


            # Adjust column widths for better readability
            for col_num in range(1, len(individualHeaders) + 1):
                column = get_column_letter(col_num)
                max_length = 0
                for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width


        for row_num in range(4, 35):
            if row_num in ruletypeRows:
                ws.cell(row=row_num, column=main_col_num, value=len(getattr(ruleAnalysis, ruletypeRows[row_num])))
            else:
                ws.cell(row=row_num, column=main_col_num, value=0)
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=3, max_col=2 + len(firewalls)):
        row_sum = sum(cell.value for cell in row)
        ws.cell(row=row[0].row, column=3 + len(firewalls), value=row_sum)

    # Create an in-memory binary stream for the Excel file
    output = BytesIO()
    
    # Save the workbook to the binary stream and seek to the beginning
    wb.save(output)
    output.seek(0)
    
    return output

